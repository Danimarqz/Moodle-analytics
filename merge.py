import openpyxl
import os

dir = os.getcwd()
dir_files = os.listdir(dir)
archivo_destino = 'Informe completo teleformación (octubre).xlsx'

# Cargar el archivo de destino
libro_destino = openpyxl.load_workbook(archivo_destino)
hoja_destino = libro_destino.active

# Iterar sobre los archivos en el directorio
for file in dir_files:
    if file.endswith('.xlsx') and file != archivo_destino:
        # Obtener el nombre de la hoja (eliminar extensión .xlsx)
        hoja_nombre = os.path.splitext(file)[0]
        # Cargar el archivo de origen
        libro_origen = openpyxl.load_workbook(file, data_only=True)
        # Obtener la primera hoja del archivo de origen
        hoja_origen = libro_origen.active
        print(libro_destino[hoja_nombre])
        # Copiar datos de la hoja de origen a la hoja de destino con el mismo nombre
        for row in hoja_origen.iter_rows(min_row=9, values_only=True):
            hoja_destino = libro_destino[hoja_nombre]
            hoja_destino.append(row)

# Guardar el archivo de destino con los datos copiados
libro_destino.save(archivo_destino)